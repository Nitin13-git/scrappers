import openpyxl 
path = "C:/Users/lenovo/Downloads/186.xlsx"
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active 

row = sheet_obj.max_row
column = sheet_obj.max_column
for i in range(1, column + 1): 
    cell_obj = sheet_obj.cell(row = 246, column = i) 
    print(cell_obj.value, end = " ")

import pandas as pd 
import numpy as np
 


